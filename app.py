#!/usr/bin/env python3
"""
app.py - Webgrensesnitt for leilighetsjakt

Start:  python3 app.py
Aapne: http://localhost:5000
"""
from __future__ import annotations

import io
import json
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import requests as http_requests
from flask import Flask, redirect, render_template_string, request, send_file, url_for

import scrape_finn
import scrape_visning
import sync_favorites

_DEFAULT_EMAIL = ""
_DEFAULT_FAVORITES_URL = ""

app = Flask(__name__)

DATA_FILE = Path(__file__).parent / "apartments.json"

# (json_key, column_header, scrape_finn data-dict key or None)
# None means the value is set manually (finnkode from URL, url from arg, hentet from now())
FIELDS = [
    ("finnkode",      "FINN-kode",      None),
    ("solgt",         "Solgt",          "Solgt"),
    ("visning",       "Visning",        "Visning"),
    ("adresse",       "Adresse",        "Adresse"),
    ("prisantydning", "Prisantydning",  "Prisantydning"),
    ("totalpris",     "Totalpris",      "Totalpris"),
    ("fellesgjeld",   "Fellesgjeld",    "Fellesgjeld"),
    ("felleskost",    "Felleskost/mnd", "Felleskost/mnd"),
    ("omkostninger",  "Omkostninger",   "Omkostninger"),
    ("bra_i",         "BRA-i",          "BRA-i"),
    ("bra",           "BRA",            "BRA"),
    ("rom",           "Rom",            "Rom"),
    ("soverom",       "Soverom",        "Soverom"),
    ("etasje",        "Etasje",         "Etasje"),
    ("byggeaar",      "Bygge\u00e5r",   "Bygge\u00e5r"),
    ("boligtype",     "Boligtype",      "Boligtype"),
    ("eieform",       "Eieform",        "Eieform"),
    ("tomteareal",    "Tomteareal",     "Tomteareal"),
    ("balkong",       "Balkong",        "Balkong"),
    ("tg3_antall",         "TG3",          "TG3 antall"),
    ("tg2_antall",         "TG2",          "TG2 antall"),
    ("tg1_antall",         "TG1",          "TG1 antall"),
    ("hoydepunkter_antall","Høydepunkter", "Høydepunkter antall"),
    ("risikoer_antall",    "Risikoer",     "Risikoer antall"),
    ("url",           "URL",            None),
    ("salgsoppgave",  "Salgsoppgave",   "Salgsoppgave"),
    ("hentet",        "Hentet",         None),
]

# Map display label -> json_key (used for Excel import)
LABEL_TO_KEY = {label: key for key, label, _ in FIELDS}

# ---------------------------------------------------------------------------
# Data helpers
# ---------------------------------------------------------------------------

def load_data() -> list[dict]:
    if DATA_FILE.exists():
        return json.loads(DATA_FILE.read_text(encoding="utf-8"))
    return []


def save_data(apartments: list[dict]) -> None:
    DATA_FILE.write_text(
        json.dumps(apartments, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def apt_from_scrape(scrape_data: dict, finnkode: str, url: str) -> dict:
    apt: dict = {
        "finnkode": finnkode,
        "url": url,
        "hentet": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    for key, _label, scrape_key in FIELDS:
        if scrape_key and scrape_key in scrape_data:
            apt[key] = scrape_data[scrape_key]
    return apt


def merge(apartments: list[dict], new_apt: dict) -> tuple[list[dict], str]:
    """Overwrite existing entry with same finnkode, or append. Returns (list, action)."""
    for i, apt in enumerate(apartments):
        if apt.get("finnkode") == new_apt.get("finnkode"):
            apartments[i] = {**apt, **{k: v for k, v in new_apt.items() if v}}
            return apartments, "Oppdatert"
    apartments.append(new_apt)
    return apartments, "Lagt til"


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def build_excel(apartments: list[dict]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leiligheter"

    headers = [label for _, label, _ in FIELDS]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # Data rows
    url_col   = next(i + 1 for i, (k, _, _) in enumerate(FIELDS) if k == "url")
    salg_col  = next(i + 1 for i, (k, _, _) in enumerate(FIELDS) if k == "salgsoppgave")

    for apt in apartments:
        row_vals = [apt.get(key, "") for key, _, _ in FIELDS]
        ws.append(row_vals)
        r = ws.max_row

        # Make URL and Salgsoppgave clickable hyperlinks
        url_val = apt.get("url", "")
        if url_val:
            cell = ws.cell(row=r, column=url_col)
            cell.value = "Annonse"
            cell.hyperlink = url_val
            cell.font = Font(color="1D4ED8", underline="single")

        salg_val = apt.get("salgsoppgave", "")
        if salg_val:
            cell = ws.cell(row=r, column=salg_col)
            cell.value = "Salgsoppgave"
            cell.hyperlink = salg_val
            cell.font = Font(color="1D4ED8", underline="single")

    # Auto-width columns
    for col_idx in range(1, len(FIELDS) + 1):
        letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=8,
        )
        ws.column_dimensions[letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Excel import
# ---------------------------------------------------------------------------

def import_excel(file_obj) -> tuple[int, int]:
    """Read uploaded xlsx, merge into apartments.json. Returns (added, updated)."""
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    apartments = load_data()
    added = updated = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        apt: dict = {}
        for col_label, val in zip(headers, row):
            key = LABEL_TO_KEY.get(col_label)
            if key:
                apt[key] = str(val).strip() if val not in (None, "") else ""
        if not apt.get("finnkode"):
            continue
        apartments, action = merge(apartments, apt)
        if action == "Lagt til":
            added += 1
        else:
            updated += 1

    save_data(apartments)
    return added, updated


# ---------------------------------------------------------------------------
# HTML template
# ---------------------------------------------------------------------------

TEMPLATE = """<!doctype html>
<html lang="nb">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Leilighetsjakt</title>
  <style>
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
    body { font-family: system-ui, -apple-system, sans-serif; background: #f8fafc; color: #1e293b; font-size: 14px; }

    .topbar {
      background: #1D4ED8; color: #fff; padding: 0.75rem 1.25rem;
      display: flex; align-items: center; gap: 1rem; flex-wrap: wrap;
    }
    .topbar h1 { font-size: 1.1rem; font-weight: 700; margin-right: auto; }

    .url-form { display: flex; gap: 0.4rem; flex: 1 1 400px; }
    .url-form input {
      flex: 1; padding: 0.45rem 0.7rem; border: none; border-radius: 5px;
      font-size: 0.9rem; min-width: 0;
    }
    .url-form input:focus { outline: 2px solid #93c5fd; }

    .btn {
      padding: 0.45rem 0.9rem; border-radius: 5px; border: none;
      font-size: 0.88rem; font-weight: 600; cursor: pointer; white-space: nowrap;
      text-decoration: none; display: inline-block;
    }
    .btn-blue   { background: #fff; color: #1D4ED8; }
    .btn-blue:hover { background: #dbeafe; }
    .btn-outline { background: transparent; color: #fff; border: 1px solid rgba(255,255,255,.5); }
    .btn-outline:hover { background: rgba(255,255,255,.1); }

    .import-label {
      padding: 0.45rem 0.9rem; border-radius: 5px; border: 1px solid rgba(255,255,255,.5);
      font-size: 0.88rem; font-weight: 600; cursor: pointer; color: #fff; white-space: nowrap;
    }
    .import-label:hover { background: rgba(255,255,255,.1); }

    .alert {
      margin: 0.6rem 1rem; padding: 0.6rem 1rem;
      border-radius: 5px; font-size: 0.9rem;
    }
    .alert-ok  { background: #dcfce7; color: #166534; border: 1px solid #86efac; }
    .alert-err { background: #fee2e2; color: #991b1b; border: 1px solid #fca5a5; }

    .count { padding: 0.4rem 1rem; font-size: 0.82rem; color: #64748b; }

    .table-wrap { overflow-x: auto; margin: 0 0 2rem; }

    table {
      border-collapse: collapse; white-space: nowrap;
      width: max-content; min-width: 100%;
    }
    thead tr { background: #1e293b; color: #fff; }
    th {
      padding: 0.5rem 0.75rem; font-size: 0.82rem; font-weight: 600;
      letter-spacing: .03em; text-align: left;
      position: sticky; top: 0; z-index: 1;
      background: #1e293b;
    }
    tbody tr:nth-child(even) { background: #f1f5f9; }
    tbody tr:hover { background: #e0f2fe; }
    tbody tr.solgt-row td { color: #94a3b8; }
    tbody tr.solgt-row { background: #fef2f2 !important; }
    .solgt-badge { background: #fef08a; color: #92400e; border-radius: 4px; padding: 1px 6px; font-size: 0.75rem; font-weight: 700; }
    td {
      padding: 0.4rem 0.75rem; font-size: 0.88rem;
      border-bottom: 1px solid #e2e8f0; max-width: 280px;
      overflow: hidden; text-overflow: ellipsis;
    }
    td a { color: #1D4ED8; text-decoration: none; }
    td a:hover { text-decoration: underline; }

    .del-btn {
      background: none; border: none; cursor: pointer;
      color: #94a3b8; font-size: 0.85rem; padding: 0 0.25rem;
    }
    .del-btn:hover { color: #dc2626; }
    .tg-fetch-btn {
      font-size: 0.75rem; color: #6366f1; text-decoration: none;
      padding: 0 0.2rem; white-space: nowrap;
    }
    .tg-fetch-btn:hover { text-decoration: underline; }

    .tg-badge {
      display: inline-block; padding: 1px 7px; border-radius: 4px;
      font-size: 0.78rem; font-weight: 700; cursor: pointer;
    }
    .tg-badge:hover { opacity: .8; }
    .tg-tg3 { background: #fee2e2; color: #991b1b; }
    .tg-tg2 { background: #fef3c7; color: #92400e; }
    .tg-tg1 { background: #dcfce7; color: #166534; }
    .risk-badge { background: #fce7f3; color: #9d174d; display: inline-block; padding: 1px 7px; border-radius: 4px; font-size: 0.78rem; font-weight: 700; cursor: pointer; }
    .risk-badge:hover { opacity: .8; }
    .highlight-badge { background: #ede9fe; color: #5b21b6; display: inline-block; padding: 1px 7px; border-radius: 4px; font-size: 0.78rem; font-weight: 700; cursor: pointer; }
    .highlight-badge:hover { opacity: .8; }
    .list-item-risk { border-left: 3px solid #ec4899; background: #fdf2f8; padding: 0.5rem 0.75rem; margin-bottom: 0.65rem; border-radius: 0 5px 5px 0; }
    .list-item-highlight { border-left: 3px solid #8b5cf6; background: #f5f3ff; padding: 0.5rem 0.75rem; margin-bottom: 0.65rem; border-radius: 0 5px 5px 0; }
    .list-item-risk h4, .list-item-highlight h4 { font-size: 0.88rem; font-weight: 600; margin-bottom: 0.3rem; }
    .list-item-risk p, .list-item-highlight p { font-size: 0.82rem; color: #475569; margin: 0.15rem 0; }

    /* TG modal */
    .modal-overlay {
      display: none; position: fixed; inset: 0;
      background: rgba(0,0,0,.45); z-index: 200;
      justify-content: center; align-items: center;
    }
    .modal-overlay.active { display: flex; }
    .tg-modal {
      background: #fff; border-radius: 10px; padding: 1.25rem 1.5rem;
      max-width: 620px; width: 94%; max-height: 80vh; overflow-y: auto;
      box-shadow: 0 8px 32px rgba(0,0,0,.25);
    }
    .tg-modal-header {
      display: flex; justify-content: space-between;
      align-items: center; margin-bottom: 1rem;
    }
    .tg-modal-header h3 { font-size: 1rem; font-weight: 700; }
    .modal-close-btn {
      background: none; border: none; cursor: pointer;
      font-size: 1.1rem; color: #64748b; padding: 0 0.25rem;
    }
    .modal-close-btn:hover { color: #0f172a; }
    .tg-item {
      border-left: 3px solid #cbd5e1; padding: 0.5rem 0.75rem;
      margin-bottom: 0.65rem; border-radius: 0 5px 5px 0;
    }
    .tg-item.TG3 { border-color: #dc2626; background: #fef2f2; }
    .tg-item.TG2 { border-color: #f59e0b; background: #fffbeb; }
    .tg-item.TG1 { border-color: #22c55e; background: #f0fdf4; }
    .tg-item h4 { font-size: 0.88rem; font-weight: 600; margin-bottom: 0.3rem; }
    .tg-item p { font-size: 0.82rem; color: #475569; margin: 0.15rem 0; }
    .tg-item p strong { color: #334155; }

    .empty { padding: 2rem; text-align: center; color: #94a3b8; font-style: italic; }
  </style>
</head>
<body>

<div class="topbar">
  <h1>🏠 Leilighetsjakt</h1>
  <form class="url-form" method="post" action="/process">
    <input type="url" name="url" required placeholder="Lim inn finn.no lenke her…" value="{{ prefill or '' }}">
    <button class="btn btn-blue" type="submit">+ Hent</button>
  </form>
  <a class="btn btn-outline" href="/export">&#8595; Excel</a>
  <form method="post" action="/import" enctype="multipart/form-data" id="import-form">
    <label class="import-label">
      &#8593; Importer Excel
      <input type="file" name="file" accept=".xlsx" style="display:none"
             onchange="document.getElementById('import-form').submit()">
    </label>
  </form>
  <a class="btn btn-outline" href="/sync">&#8635; Sync favoritter</a>
  <a class="btn btn-outline" href="/refresh-all" onclick="return confirm('Oppdater alle leiligheter fra finn.no?')">&#8635; Oppdater alle</a>
  <a class="btn btn-outline" href="/hent-tg-alle" onclick="return confirm('Hent TG-data fra visning.ai for alle leiligheter? Dette kan ta litt tid.')">&#11015; Hent TG alle</a>
  <a class="btn btn-outline" href="/kart">&#128506; Kart</a>
</div>

{% if message %}
<div class="alert {{ 'alert-ok' if ok else 'alert-err' }}">{{ message }}</div>
{% endif %}

<div class="count">{{ apartments|length }} leilighet{{ 'er' if apartments|length != 1 else '' }} lagret</div>

<div class="table-wrap">
{% if apartments %}
<table>
  <thead>
    <tr>
      <th></th>
      {% for _, label, _ in fields %}<th>{{ label }}</th>{% endfor %}
    </tr>
  </thead>
  <tbody>
    {% for apt in apartments %}
    <tr{% if apt.get('solgt') %} class="solgt-row"{% endif %}>
      <td>
        <form method="post" action="/delete/{{ apt.get('finnkode','') }}" style="display:inline">
          <button class="del-btn" title="Slett">✕</button>
        </form>
        <a class="tg-fetch-btn" href="/hent-tg/{{ apt.get('finnkode','') }}" title="Hent TG-data fra visning.ai">⬇TG</a>
      </td>
      {% for key, label, _ in fields %}
      <td>
        {% if key == 'url' and apt.get(key) %}
          <a href="{{ apt[key] }}" target="_blank" rel="noopener">Annonse ↗</a>
        {% elif key == 'salgsoppgave' and apt.get(key) %}
          <a href="{{ apt[key] }}" target="_blank" rel="noopener">Salgsoppgave ↗</a>
        {% elif key == 'solgt' and apt.get(key) %}
          <span class="solgt-badge">Solgt</span>
        {% elif key in ('tg3_antall', 'tg2_antall', 'tg1_antall') %}
          {% set grade = 'TG3' if key == 'tg3_antall' else ('TG2' if key == 'tg2_antall' else 'TG1') %}
          {% if apt.get(key) %}
          <span class="tg-badge tg-{{ grade | lower }}"
                onclick="showTG('{{ apt.get('finnkode','') }}', '{{ grade }}')">
            {{ apt.get(key, '') }}
          </span>
          {% endif %}
        {% elif key == 'hoydepunkter_antall' %}
          {% if apt.get(key) %}
          <span class="highlight-badge" onclick="showHighlights('{{ apt.get('finnkode','') }}')">
            ✨ {{ apt.get(key, '') }}
          </span>
          {% endif %}
        {% elif key == 'risikoer_antall' %}
          {% if apt.get(key) %}
          <span class="risk-badge" onclick="showRisks('{{ apt.get('finnkode','') }}')">
            ⚠ {{ apt.get(key, '') }}
          </span>
          {% endif %}
        {% else %}
          {{ apt.get(key, '') }}
        {% endif %}
      </td>
      {% endfor %}
    </tr>
    {% endfor %}
  </tbody>
</table>
{% else %}
<div class="empty">Ingen leiligheter enda – lim inn en finn.no lenke ovenfor.</div>
{% endif %}
</div>

<!-- TG detail modal -->
<div class="modal-overlay" id="tg-modal" onclick="if(event.target===this)this.classList.remove('active')">
  <div class="tg-modal">
    <div class="tg-modal-header">
      <h3 id="tg-modal-title">TG detaljer</h3>
      <button class="modal-close-btn" onclick="document.getElementById('tg-modal').classList.remove('active')">✕</button>
    </div>
    <div id="tg-modal-body"></div>
  </div>
</div>

<script>
var TG_DATA = {
{% for apt in apartments %}
  {{ apt.get('finnkode','') | tojson }}: {{ apt.get('tg_items', '[]') | safe }},
{% endfor %}
};
var HIGHLIGHTS_DATA = {
{% for apt in apartments %}
  {{ apt.get('finnkode','') | tojson }}: {{ apt.get('hoydepunkter_items', '[]') | safe }},
{% endfor %}
};
var RISKS_DATA = {
{% for apt in apartments %}
  {{ apt.get('finnkode','') | tojson }}: {{ apt.get('risikoer_items', '[]') | safe }},
{% endfor %}
};

function _openModal(title, bodyHtml) {
  document.getElementById('tg-modal-title').textContent = title;
  document.getElementById('tg-modal-body').innerHTML = bodyHtml;
  document.getElementById('tg-modal').classList.add('active');
}

function showHighlights(finnkode) {
  var items = HIGHLIGHTS_DATA[finnkode] || [];
  var html = '';
  if (!items.length) { html = '<p style="color:#64748b">Ingen høydepunkter.</p>'; }
  else { items.forEach(function(it) {
    html += '<div class="list-item-highlight"><h4>' + (it.emoji ? it.emoji + ' ' : '') + (it.header || '') + '</h4>';
    if (it.details) html += '<p>' + it.details + '</p>';
    if (it.category) html += '<p style="font-size:0.75rem;color:#7c3aed">' + it.category + '</p>';
    html += '</div>';
  }); }
  _openModal('Høydepunkter (' + items.length + ')', html);
}

function showRisks(finnkode) {
  var items = RISKS_DATA[finnkode] || [];
  var html = '';
  if (!items.length) { html = '<p style="color:#64748b">Ingen risikoer.</p>'; }
  else { items.forEach(function(it) {
    html += '<div class="list-item-risk"><h4>' + (it.emoji ? it.emoji + ' ' : '') + (it.header || '') + '</h4>';
    if (it.details) html += '<p>' + it.details + '</p>';
    if (it.question) html += '<p style="font-size:0.78rem;color:#9d174d"><em>Spørsmål: ' + it.question + '</em></p>';
    if (it.category) html += '<p style="font-size:0.75rem;color:#be185d">' + it.category + '</p>';
    html += '</div>';
  }); }
  _openModal('Risikoer (' + items.length + ')', html);
}

function showTG(finnkode, grade) {
  var items = TG_DATA[finnkode] || [];
  var filtered = grade ? items.filter(function(it){ return it.grade === grade; }) : items;
  var title = grade + ' \u2013 ' + filtered.length + ' element' + (filtered.length !== 1 ? 'er' : '');
  var html = '';
  if (!filtered.length) {
    html = '<p style="color:#64748b">Ingen elementer.</p>';
  } else {
    filtered.forEach(function(it) {
      html += '<div class="tg-item ' + (it.grade || '') + '">';
      html += '<h4>' + (it.name || '') + (it.location ? ' (' + it.location + ')' : '') + '</h4>';
      if (it.justification)      html += '<p><strong>Vurdering:</strong> ' + it.justification + '</p>';
      if (it.risk)               html += '<p><strong>Risiko:</strong> ' + it.risk + '</p>';
      if (it.recommendedActions) html += '<p><strong>Tiltak:</strong> ' + it.recommendedActions + '</p>';
      if (it.costEstimate)       html += '<p><strong>Kostnad:</strong> ' + it.costEstimate + '</p>';
      html += '</div>';
    });
  }
  _openModal(title, html);
}
</script>

</body>
</html>"""


# ---------------------------------------------------------------------------
# Sync page template
# ---------------------------------------------------------------------------

SYNC_TEMPLATE = """<!doctype html>
<html lang="nb">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Sync favoritter – Leilighetsjakt</title>
  <style>
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
    body { font-family: system-ui, -apple-system, sans-serif; background: #f8fafc; color: #1e293b; font-size: 14px; }
    .topbar { background: #1D4ED8; color: #fff; padding: 0.75rem 1.25rem; display: flex; align-items: center; gap: 1rem; }
    .topbar h1 { font-size: 1.1rem; font-weight: 700; }
    .topbar a { color: #fff; font-size: 0.88rem; margin-left: auto; text-decoration: none; opacity: .8; }
    .topbar a:hover { opacity: 1; }
    .card { background: #fff; border-radius: 10px; box-shadow: 0 1px 4px rgba(0,0,0,.1); padding: 1.5rem; margin: 1.5rem auto; max-width: 680px; }
    h2 { font-size: 1.1rem; margin-bottom: 1rem; }
    label { display: block; font-size: 0.88rem; font-weight: 600; margin-bottom: 0.3rem; color: #475569; }
    input[type="url"], input[type="text"] {
      width: 100%; padding: 0.55rem 0.8rem; border: 1px solid #cbd5e1;
      border-radius: 6px; font-size: 0.95rem; margin-bottom: 1rem;
    }
    input:focus { outline: 2px solid #3b82f6; border-color: transparent; }
    .btn { padding: 0.55rem 1.2rem; border-radius: 6px; border: none; font-size: 0.92rem; font-weight: 600; cursor: pointer; }
    .btn-primary { background: #1D4ED8; color: #fff; }
    .btn-primary:hover { background: #1e40af; }
    .btn-primary:disabled { opacity: .5; cursor: default; }
    .progress-bar-wrap { background: #e2e8f0; border-radius: 99px; height: 8px; margin: 1rem 0; overflow: hidden; }
    .progress-bar { height: 100%; background: #1D4ED8; border-radius: 99px; transition: width .4s; }
    .log-box { background: #0f172a; color: #94a3b8; border-radius: 6px; padding: 0.75rem 1rem; font-family: monospace; font-size: 0.82rem; max-height: 220px; overflow-y: auto; margin-top: 1rem; }
    .log-box p { margin: 0.1rem 0; }
    .log-box p:last-child { color: #e2e8f0; }
    .status-badge { display: inline-block; padding: 0.2rem 0.6rem; border-radius: 99px; font-size: 0.78rem; font-weight: 700; letter-spacing: .04em; }
    .badge-idle      { background: #e2e8f0; color: #64748b; }
    .badge-starting  { background: #dbeafe; color: #1e40af; }
    .badge-waiting   { background: #fef9c3; color: #92400e; }
    .badge-scraping  { background: #d1fae5; color: #065f46; }
    .badge-done      { background: #d1fae5; color: #065f46; }
    .badge-error     { background: #fee2e2; color: #991b1b; }
    .otp-section { background: #fffbeb; border: 1px solid #fcd34d; border-radius: 8px; padding: 1rem 1.25rem; margin-top: 1rem; }
    .otp-section h3 { font-size: 0.95rem; margin-bottom: 0.5rem; color: #92400e; }
    .otp-row { display: flex; gap: 0.5rem; }
    .otp-row input { flex: 1; margin: 0; letter-spacing: .2em; font-size: 1.2rem; text-align: center; }
  </style>
</head>
<body>
<div class="topbar">
  <h1>&#8635; Sync favoritter</h1>
  <a href="/">&larr; Tilbake til tabellen</a>
</div>

<div class="card">
  {% if state.status == 'idle' %}
  <h2>Synkroniser favorittlisten din</h2>
  <form method="post" action="/sync/start">
    <label>E-post (brukes til å logge inn på finn.no)</label>
    <input type="email" name="email" required
           value="{{ default_email }}"
           placeholder="din@epost.no">
    <label>Delt favorittliste-lenke fra finn.no</label>
    <input type="url" name="list_url" required
           value="{{ default_favorites_url }}"
           placeholder="https://www.finn.no/sharedfavoritelist/…">
    <button class="btn btn-primary" type="submit">Start sync</button>
  </form>

  {% else %}
  <div style="display:flex; align-items:center; gap:.75rem; margin-bottom:1rem;">
    <h2 style="margin:0">Status</h2>
    <span class="status-badge badge-{{ state.status.replace('_','-').replace('waiting-for-code','waiting') }}">
      {{ state.status.replace('_', ' ') }}
    </span>
  </div>

  <p style="color:#334155; font-size:.95rem;">{{ state.message }}</p>

  {% if state.status == 'scraping' and state.found > 0 %}
  <div class="progress-bar-wrap" style="margin-top:.75rem;">
    <div class="progress-bar" style="width: {{ (state.scraped / state.found * 100) | int }}%"></div>
  </div>
  <p style="font-size:.82rem; color:#64748b;">{{ state.scraped }} / {{ state.found }} hentet</p>
  {% endif %}

  {% if state.status == 'waiting_for_code' %}
  <div class="otp-section">
    <h3>Skriv inn koden fra e-posten</h3>
    <form method="post" action="/sync/code">
      <div class="otp-row">
        <input type="text" name="code" maxlength="6" pattern="[0-9]{6}"
               inputmode="numeric" autocomplete="one-time-code"
               placeholder="000000" autofocus required>
        <button class="btn btn-primary" type="submit">Send kode</button>
      </div>
    </form>
  </div>
  {% endif %}

  {% if state.log %}
  <div class="log-box" id="log">
    {% for line in state.log %}<p>{{ line }}</p>{% endfor %}
  </div>
  {% endif %}

  {% if state.status == 'done' %}
  <div style="margin-top:1rem; display:flex; gap:.5rem;">
    <a class="btn btn-primary" href="/" style="text-decoration:none;">Se tabellen &rarr;</a>
    <a class="btn" style="background:#e2e8f0;" href="/sync">Ny sync</a>
  </div>
  {% elif state.status == 'error' %}
  <div style="margin-top:1rem;">
    <a class="btn" style="background:#e2e8f0;" href="/sync">Provo igjen</a>
  </div>
  {% endif %}

  {% endif %}
</div>

{% if state.status not in ('idle', 'done', 'error') %}
<script>
  // Poll status every 2 seconds and refresh the page
  setTimeout(function poll() {
    fetch('/sync/status')
      .then(r => r.json())
      .then(s => {
        if (s.status !== '{{ state.status }}' || s.message !== {{ state.message | tojson }}) {
          location.reload();
        } else {
          // scroll log to bottom
          var log = document.getElementById('log');
          if (log) log.scrollTop = log.scrollHeight;
          setTimeout(poll, 2000);
        }
      })
      .catch(() => setTimeout(poll, 3000));
  }, 2000);
</script>
{% endif %}
</body>
</html>"""

# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

def _render(apartments=None, message=None, ok=True, prefill=""):
    if apartments is None:
        apartments = load_data()
    return render_template_string(
        TEMPLATE, apartments=apartments, message=message,
        ok=ok, prefill=prefill, fields=FIELDS,
    )


@app.get("/")
def index():
    return _render()


@app.post("/process")
def process():
    url = request.form.get("url", "").strip()

    try:
        finnkode = scrape_finn.extract_finnkode(url)
    except ValueError:
        return _render(message=f"Ugyldig URL – fant ingen finnkode: {url}", ok=False, prefill=url)

    try:
        soup = scrape_finn.fetch_page(url)
    except http_requests.RequestException as exc:
        return _render(message=f"Nettverksfeil: {exc}", ok=False, prefill=url)

    data = scrape_finn.scrape(soup)
    if not data:
        return _render(message="Ingen data funnet. Sjekk at URL-en er en finn.no boligannonse.", ok=False, prefill=url)

    # Persist to JSON
    apt = apt_from_scrape(data, finnkode, url)
    apartments, action = merge(load_data(), apt)
    save_data(apartments)

    # Keep apartments.md in sync
    scrape_finn.update_file(scrape_finn.format_section(data, finnkode, url), finnkode)

    label = data.get("Adresse") or data.get("Tittel", finnkode)
    return _render(apartments=apartments, message=f"✓ {action}: {label}", ok=True)


@app.get("/export")
def export_excel():
    buf = build_excel(load_data())
    filename = f"leiligheter_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/import")
def import_excel_route():
    f = request.files.get("file")
    if not f or not f.filename.lower().endswith(".xlsx"):
        return _render(message="Velg en .xlsx fil.", ok=False)
    added, updated = import_excel(f)
    return _render(message=f"✓ Importert: {added} nye, {updated} oppdatert")


@app.post("/delete/<finnkode>")
def delete(finnkode: str):
    save_data([a for a in load_data() if a.get("finnkode") != finnkode])
    return redirect(url_for("index"))


@app.get("/refresh-all")
def refresh_all():
    import scrape_finn as sf
    apartments = load_data()
    updated = 0
    errors = 0
    for i, apt in enumerate(apartments):
        url = apt.get("url", "")
        finnkode = apt.get("finnkode", "")
        if not url or not finnkode:
            continue
        try:
            soup = sf.fetch_page(url)
            data = sf.scrape(soup)
            if data:
                new_apt = apt_from_scrape(data, finnkode, url)
                apartments[i] = {**apt, **{k: v for k, v in new_apt.items() if v is not None}}
                # Explicitly clear solgt if not present in new scrape
                if "Solgt" not in data:
                    apartments[i].pop("solgt", None)
                updated += 1
        except Exception:
            errors += 1
    save_data(apartments)
    msg = f"Oppdatert {updated} leiligheter."
    if errors:
        msg += f" {errors} feil."
    return render_template_string(
        TEMPLATE, apartments=apartments, message=msg,
        ok=True, prefill="", fields=FIELDS,
    )


# ---------------------------------------------------------------------------
# visning.ai TG routes
# ---------------------------------------------------------------------------

def _apply_visning_data(apartments: list[dict], i: int, finnkode: str) -> str:
    """Fetch TG data from visning.ai and update apartments[i] in-place. Returns status message."""
    data = scrape_visning.scrape(finnkode)
    if not data:
        return f"Ingen TG-data funnet for {finnkode} på visning.ai"
    tg_items = data.pop("tg_items", None)
    hoydepunkter_items = data.pop("hoydepunkter_items", None)
    risikoer_items = data.pop("risikoer_items", None)
    for key, _, scrape_key in FIELDS:
        if scrape_key and scrape_key in data:
            apartments[i][key] = data[scrape_key]
    if tg_items is not None:
        apartments[i]["tg_items"] = tg_items
    if hoydepunkter_items is not None:
        apartments[i]["hoydepunkter_items"] = hoydepunkter_items
    if risikoer_items is not None:
        apartments[i]["risikoer_items"] = risikoer_items
    return "ok"


@app.get("/hent-tg/<finnkode>")
def hent_tg(finnkode: str):
    apartments = load_data()
    for i, apt in enumerate(apartments):
        if apt.get("finnkode") == finnkode:
            try:
                msg = _apply_visning_data(apartments, i, finnkode)
                save_data(apartments)
                if msg == "ok":
                    adresse = apt.get("adresse", finnkode)
                    return _render(apartments=apartments, message=f"✓ TG-data hentet for {adresse}", ok=True)
                return _render(apartments=apartments, message=msg, ok=False)
            except RuntimeError as exc:
                return _render(apartments=apartments, message=f"Feil: {exc}", ok=False)
    return _render(message=f"Fant ikke finnkode {finnkode}", ok=False)


@app.get("/hent-tg-alle")
def hent_tg_alle():
    apartments = load_data()
    ok_count = errors = skipped = 0
    for i, apt in enumerate(apartments):
        finnkode = apt.get("finnkode", "")
        if not finnkode:
            continue
        try:
            msg = _apply_visning_data(apartments, i, finnkode)
            if msg == "ok":
                ok_count += 1
            else:
                skipped += 1
        except RuntimeError:
            errors += 1
    save_data(apartments)
    parts = [f"TG-data hentet for {ok_count} leiligheter."]
    if skipped:
        parts.append(f"{skipped} uten data på visning.ai.")
    if errors:
        parts.append(f"{errors} feil.")
    return render_template_string(
        TEMPLATE, apartments=apartments, message=" ".join(parts),
        ok=True, prefill="", fields=FIELDS,
    )


# ---------------------------------------------------------------------------
# Geocoding
# ---------------------------------------------------------------------------

_NOMINATIM_HEADERS = {
    "User-Agent": "leilighetsjakt/1.0 (personal apartment search tool)",
    "Accept-Language": "nb,no,en",
}


def geocode_address(address: str) -> tuple[float, float] | None:
    """Geocode a Norwegian address via Nominatim. Returns (lat, lon) or None."""
    import re
    # Strip parenthetical notes and ' - <description>' suffixes that confuse Nominatim
    cleaned = re.sub(r"\s*\([^)]*\)", "", address).strip()
    cleaned = re.split(r"\s+-\s+", cleaned)[0].strip()
    
    def try_geocode(query: str) -> tuple[float, float] | None:
        """Helper to attempt geocoding with a query string."""
        try:
            r = http_requests.get(
                "https://nominatim.openstreetmap.org/search",
                params={"q": query, "format": "json", "limit": 1, "countrycodes": "no"},
                headers=_NOMINATIM_HEADERS,
                timeout=10,
            )
            r.raise_for_status()
            results = r.json()
            if results:
                return float(results[0]["lat"]), float(results[0]["lon"])
        except Exception:
            pass
        return None
    
    # Try with full address first
    result = try_geocode(cleaned)
    if result:
        return result
    
    # Try removing building numbers with letters (e.g., "3C", "1A") that confuse Nominatim
    # These appear after the street name and number, like "Olaf Ryes plass 3C" -> "Olaf Ryes plass"
    no_letter_suffix = re.sub(r"\s+\d+[A-Z]\b", "", cleaned, flags=re.IGNORECASE).strip()
    if no_letter_suffix != cleaned:
        result = try_geocode(no_letter_suffix)
        if result:
            return result
    
    # Try removing postal code (which is often wrong/incomplete)
    simplified = re.sub(r",?\s*\d{4}\b", "", no_letter_suffix).strip()
    simplified = re.sub(r",\s+$", "", simplified).strip()
    if simplified != no_letter_suffix:
        result = try_geocode(simplified)
        if result:
            return result
    
    # Try just street address without number
    street_only = re.sub(r"\s+\d+.*?(?:,|$)", ",", cleaned).strip()
    street_only = re.sub(r",\s+$", "", street_only).strip()
    if street_only and street_only != cleaned and street_only != no_letter_suffix:
        result = try_geocode(street_only)
        if result:
            return result
    
    return None


# ---------------------------------------------------------------------------
# Map template
# ---------------------------------------------------------------------------

MAP_TEMPLATE = """<!doctype html>
<html lang="nb">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Leilighetskart</title>
  <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
  <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
  <style>
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
    body { font-family: system-ui, -apple-system, sans-serif; background: #f8fafc; color: #1e293b; display: flex; flex-direction: column; height: 100vh; }
    .topbar {
      background: #1D4ED8; color: #fff; padding: 0.65rem 1.25rem;
      display: flex; align-items: center; gap: 1rem; flex-shrink: 0;
    }
    .topbar h1 { font-size: 1.05rem; font-weight: 700; }
    .btn { padding: 0.4rem 0.85rem; border-radius: 5px; border: none; font-size: 0.85rem; font-weight: 600; cursor: pointer; text-decoration: none; display: inline-block; }
    .btn-outline { background: transparent; color: #fff; border: 1px solid rgba(255,255,255,.5); }
    .btn-outline:hover { background: rgba(255,255,255,.1); }
    .status { font-size: 0.82rem; color: #93c5fd; margin-left: auto; }
    #map { flex: 1; }
    .popup-title { font-weight: 700; font-size: 0.95rem; margin-bottom: 4px; }
    .popup-row { font-size: 0.82rem; color: #334155; margin: 2px 0; }
    .popup-link { font-size: 0.82rem; color: #1D4ED8; text-decoration: none; }
    .popup-link:hover { text-decoration: underline; }
    .popup-solgt { background: #fef08a; color: #92400e; border-radius: 3px; padding: 1px 5px; font-size: 0.75rem; font-weight: 700; }
    .popup-visning { background: #dcfce7; color: #166534; border-radius: 3px; padding: 1px 5px; font-size: 0.75rem; font-weight: 600; }
    .legend {
      position: absolute; bottom: 24px; right: 12px; z-index: 1000;
      background: #fff; border-radius: 8px; padding: 0.6rem 0.85rem;
      box-shadow: 0 2px 8px rgba(0,0,0,.2); font-size: 0.8rem; line-height: 1.8;
    }
    .legend-dot {
      display: inline-block; width: 12px; height: 12px; border-radius: 50%;
      border: 2px solid #fff; box-shadow: 0 1px 3px rgba(0,0,0,.35); margin-right: 6px; vertical-align: middle;
    }
  </style>
</head>
<body>
<div class="topbar">
  <h1>&#128506; Leilighetskart</h1>
  <a class="btn btn-outline" href="/">&larr; Tilbake til tabell</a>
  <a class="btn btn-outline" href="/geocode-alle" id="geocode-alle-btn">&#128269; Geokoder alle adresser</a>
  <span class="status" id="status"></span>
</div>
<div id="map"></div>
<div class="legend">
  <div><span class="legend-dot" style="background:#1D4ED8"></span>Tilgjengelig</div>
  <div><span class="legend-dot" style="background:#dc2626"></span>Har visning</div>
  <div><span class="legend-dot" style="background:#94a3b8"></span>Solgt</div>
</div>
<script>
var APTS = {{ apts_json | safe }};

var map = L.map('map').setView([59.91, 10.75], 12);
L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', {
  attribution: '&copy; <a href="https://www.openstreetmap.org/copyright">OpenStreetMap</a> contributors',
  maxZoom: 19,
}).addTo(map);

var bounds = [];
var geocodeQueue = [];

function makePopup(apt) {
  var html = '<div class="popup-title">' + (apt.adresse || apt.finnkode) + '</div>';
  if (apt.solgt) html += ' <span class="popup-solgt">Solgt</span>';
  if (apt.visning) html += '<div><span class="popup-visning">&#128197; ' + apt.visning + '</span></div>';
  if (apt.prisantydning) html += '<div class="popup-row">Prisantydning: <strong>' + apt.prisantydning + '</strong></div>';
  if (apt.totalpris)    html += '<div class="popup-row">Totalpris: ' + apt.totalpris + '</div>';
  if (apt.bra_i)        html += '<div class="popup-row">BRA-i: ' + apt.bra_i + '</div>';
  if (apt.soverom)      html += '<div class="popup-row">Soverom: ' + apt.soverom + '</div>';
  if (apt.etasje)       html += '<div class="popup-row">Etasje: ' + apt.etasje + '</div>';
  if (apt.url)          html += '<div style="margin-top:6px"><a class="popup-link" href="' + apt.url + '" target="_blank" rel="noopener">Annonse &#8599;</a></div>';
  return html;
}

function addMarker(apt, lat, lon) {
  var color = apt.solgt ? '#94a3b8' : (apt.visning ? '#dc2626' : '#1D4ED8');
  var icon = L.divIcon({
    html: '<div style="width:14px;height:14px;border-radius:50%;background:' + color + ';border:2px solid #fff;box-shadow:0 1px 4px rgba(0,0,0,.4)"></div>',
    className: '', iconAnchor: [7, 7],
  });
  var marker = L.marker([lat, lon], {icon: icon}).addTo(map);
  marker.bindPopup(makePopup(apt), {maxWidth: 280});
  bounds.push([lat, lon]);
}

// Place apartments that already have coordinates
APTS.forEach(function(apt) {
  if (apt.lat && apt.lon) {
    addMarker(apt, apt.lat, apt.lon);
  } else if (apt.adresse) {
    geocodeQueue.push(apt);
  }
});

if (bounds.length) map.fitBounds(bounds, {padding: [40, 40]});

// Geocode missing addresses one by one (Nominatim rate limit: 1/s)
var geocodeIndex = 0;
function geocodeNext() {
  if (geocodeIndex >= geocodeQueue.length) {
    document.getElementById('status').textContent = '';
    return;
  }
  var apt = geocodeQueue[geocodeIndex++];
  document.getElementById('status').textContent =
    'Geokoder ' + geocodeIndex + '/' + geocodeQueue.length + ': ' + apt.adresse;
  fetch('/geocode/' + encodeURIComponent(apt.finnkode))
    .then(function(r){ return r.json(); })
    .then(function(d){
      if (d.lat && d.lon) {
        apt.lat = d.lat; apt.lon = d.lon;
        addMarker(apt, d.lat, d.lon);
        if (bounds.length === 1) map.setView([d.lat, d.lon], 13);
        else map.fitBounds(bounds, {padding: [40, 40]});
      }
      setTimeout(geocodeNext, 1000);
    })
    .catch(function(){ setTimeout(geocodeNext, 1000); });
}

// Skip geocoding if all already cached
if (geocodeQueue.length > 0) {
  var geocodeAllBtn = document.getElementById('geocode-alle-btn');
  if (geocodeAllBtn) {
    geocodeAllBtn.textContent = '\u23F3 Geokoder alle (' + geocodeQueue.length + ' mangler)...';
    geocodeAllBtn.onclick = function(e) {
      e.preventDefault();
      geocodeAllBtn.textContent = '\u23F3 Laster...';
      window.location = '/geocode-alle';
    };
  }
  geocodeNext();
} else {
  var geocodeAllBtn = document.getElementById('geocode-alle-btn');
  if (geocodeAllBtn) geocodeAllBtn.style.display = 'none';
}
</script>
</body>
</html>"""


@app.get("/kart")
def kart():
    apartments = load_data()
    # Build a lightweight list for the map (only fields needed for popups)
    apt_keys = ["finnkode", "adresse", "solgt", "visning", "prisantydning",
                "totalpris", "bra_i", "soverom", "etasje", "url", "lat", "lon"]
    apts = [{k: apt.get(k, "") for k in apt_keys} for apt in apartments]
    return render_template_string(MAP_TEMPLATE, apts_json=json.dumps(apts, ensure_ascii=False))


@app.get("/geocode/<finnkode>")
def geocode_apt(finnkode: str):
    """Geocode one apartment by address, store result, return {lat, lon}."""
    apartments = load_data()
    for i, apt in enumerate(apartments):
        if apt.get("finnkode") == finnkode:
            # Return cached coords if already set
            if apt.get("lat") and apt.get("lon"):
                return {"lat": apt["lat"], "lon": apt["lon"]}
            address = apt.get("adresse", "")
            if not address:
                return {"error": "no address"}, 404
            coords = geocode_address(address)
            if coords:
                apartments[i]["lat"] = coords[0]
                apartments[i]["lon"] = coords[1]
                save_data(apartments)
                return {"lat": coords[0], "lon": coords[1]}
            return {"error": "not found"}, 404
    return {"error": "unknown finnkode"}, 404


@app.get("/geocode-alle")
def geocode_alle():
    """Server-side batch geocode all apartments missing coordinates. Respects Nominatim 1/s limit."""
    import time
    apartments = load_data()
    updated = 0
    for i, apt in enumerate(apartments):
        if apt.get("lat") and apt.get("lon"):
            continue
        address = apt.get("adresse", "")
        if not address:
            continue
        coords = geocode_address(address)
        if coords:
            apartments[i]["lat"] = coords[0]
            apartments[i]["lon"] = coords[1]
            updated += 1
        time.sleep(1.0)  # Nominatim rate limit
    save_data(apartments)
    return redirect(url_for("kart"))


# ---------------------------------------------------------------------------
# Sync favorites routes
# ---------------------------------------------------------------------------

def _sync_render(state=None):
    if state is None:
        state = sync_favorites.get_state()
    return render_template_string(
        SYNC_TEMPLATE, state=state,
        default_email=_DEFAULT_EMAIL,
        default_favorites_url=_DEFAULT_FAVORITES_URL,
    )


def _on_apt_scraped(finnkode: str, url: str) -> str:
    """Callback from the sync thread: scrape one apartment and save it."""
    import scrape_finn as sf
    soup = sf.fetch_page(url)
    data = sf.scrape(soup)
    if not data:
        raise ValueError("Ingen data funnet")
    apt = apt_from_scrape(data, finnkode, url)
    apartments, action = merge(load_data(), apt)
    save_data(apartments)
    sf.update_file(sf.format_section(data, finnkode, url), finnkode)
    return action


@app.get("/sync")
def sync_page():
    # Reset to idle if previous run is finished
    state = sync_favorites.get_state()
    if state["status"] in ("done", "error"):
        return _sync_render({"status": "idle", "message": "", "log": [],
                             "found": 0, "scraped": 0, "added": 0, "updated": 0})
    return _sync_render(state)


@app.post("/sync/start")
def sync_start():
    if sync_favorites.is_running():
        return _sync_render()
    list_url = request.form.get("list_url", "").strip()
    email = request.form.get("email", "").strip()
    if not list_url:
        return _sync_render({"status": "error", "message": "Mangler URL.",
                             "log": [], "found": 0, "scraped": 0, "added": 0, "updated": 0})
    if not email:
        return _sync_render({"status": "error", "message": "Mangler e-post.",
                             "log": [], "found": 0, "scraped": 0, "added": 0, "updated": 0})
    sync_favorites.start_sync(list_url, email, _on_apt_scraped)
    return redirect(url_for("sync_wait"))


@app.get("/sync/wait")
def sync_wait():
    return _sync_render()


@app.get("/sync/status")
def sync_status():
    from flask import jsonify
    return jsonify(sync_favorites.get_state())


@app.post("/sync/code")
def sync_code():
    code = request.form.get("code", "").strip()
    if code:
        sync_favorites.submit_code(code)
    return redirect(url_for("sync_wait"))


# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print("Starter webserver paa http://localhost:5000")
    app.run(debug=False, port=5000)
